from __future__ import annotations

import csv
from collections.abc import Iterable, Sequence
from datetime import date, datetime, timezone
from io import BytesIO, StringIO
from typing import Any

from fastapi import Response
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.core.config import get_app_timezone, get_settings

CSV_MEDIA_TYPE = "text/csv; charset=utf-8"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def current_app_date() -> date:
    settings = get_settings()
    return datetime.now(timezone.utc).astimezone(get_app_timezone(settings.app_timezone)).date()


def _sanitize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if not isinstance(value, str):
        return value

    cleaned = value.replace("\x00", "")
    if cleaned.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{cleaned}"
    return cleaned


def _sanitize_rows(rows: Iterable[Sequence[Any]]) -> list[list[Any]]:
    return [[_sanitize_cell(value) for value in row] for row in rows]


def csv_export_response(
    *,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    filename: str,
) -> Response:
    buffer = StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow(headers)
    writer.writerows(_sanitize_rows(rows))
    content = ("\ufeff" + buffer.getvalue()).encode("utf-8")
    return Response(
        content=content,
        media_type=CSV_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "private, no-store",
        },
    )


def xlsx_export_response(
    *,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    filename: str,
    worksheet_title: str,
) -> Response:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_title[:31]
    worksheet.append(list(headers))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    sanitized_rows = _sanitize_rows(rows)
    for row in sanitized_rows:
        worksheet.append(row)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(sanitized_rows) + 1)}"
    for index, header in enumerate(headers, start=1):
        values = [header, *(row[index - 1] for row in sanitized_rows)]
        width = min(max(max((len(str(value)) for value in values), default=0) + 2, 12), 42)
        worksheet.column_dimensions[get_column_letter(index)].width = width
        if header == "usage_percentage":
            for cell in worksheet[get_column_letter(index)][1:]:
                cell.number_format = "0.00"

    buffer = BytesIO()
    workbook.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "private, no-store",
        },
    )
