"""Integration coverage for user-scoped CSV and XLSX exports."""
from __future__ import annotations

import csv
from datetime import date
from io import BytesIO, StringIO

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

from app.core.exports import XlsxWorksheet, current_app_date, xlsx_workbook_response
from app.modules.finances import service as finances_service

pytestmark = pytest.mark.asyncio


async def _register_and_login(client: AsyncClient, email: str) -> str:
    password = "correcthorse"
    await client.post("/api/v1/auth/register", json={"email": email, "password": password})
    response = await client.post("/api/v1/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200, response.text
    return response.json()["access_token"]


def _headers(token: str) -> dict[str, str]:
    return {"authorization": f"Bearer {token}"}


async def _create_account_and_category(client: AsyncClient, token: str) -> tuple[str, str]:
    headers = _headers(token)
    account = await client.post(
        "/api/v1/finances/accounts",
        headers=headers,
        json={"name": "Primary account", "type": "checking", "initial_balance": 0},
    )
    category = await client.post(
        "/api/v1/finances/categories",
        headers=headers,
        json={"name": "=Formula category", "type": "expense"},
    )
    assert account.status_code == 201
    assert category.status_code == 201
    return account.json()["id"], category.json()["id"]


def _csv_rows(response) -> list[list[str]]:
    assert response.content.startswith(b"\xef\xbb\xbf")
    return list(csv.reader(StringIO(response.content.decode("utf-8-sig"))))


async def test_export_endpoints_require_authentication(client: AsyncClient) -> None:
    for path in (
        "/api/v1/finances/exports/transactions.csv",
        "/api/v1/finances/exports/transactions.xlsx",
        "/api/v1/finances/exports/monthly-budgets.csv",
        "/api/v1/finances/exports/monthly-budgets.xlsx",
        "/api/v1/finances/reports/monthly.xlsx",
        "/api/v1/savings/exports/goals.csv",
        "/api/v1/savings/exports/goals.xlsx",
    ):
        response = await client.get(path)
        assert response.status_code == 401


async def test_multi_sheet_xlsx_sanitizes_cells_and_formats_percentages() -> None:
    response = xlsx_workbook_response(
        worksheets=[
            XlsxWorksheet(
                title="Summary",
                headers=("name", "savings_rate"),
                rows=[["\x00=Unsafe", 12.5]],
                percentage_headers=frozenset({"savings_rate"}),
            )
        ],
        filename="test.xlsx",
    )
    workbook = load_workbook(BytesIO(response.body))
    worksheet = workbook.active
    assert worksheet["A2"].value == "'=Unsafe"
    assert worksheet["B2"].value == 12.5
    assert worksheet["B2"].number_format == "0.00"
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:B2"


async def test_monthly_report_xlsx_export_reuses_report_data_and_is_scoped(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(finances_service, "current_app_date", lambda: date(2026, 7, 22))
    owner = await _register_and_login(client, "report-export-owner@example.com")
    other = await _register_and_login(client, "report-export-other@example.com")
    owner_headers = _headers(owner)
    account_id, category_id = await _create_account_and_category(client, owner)
    income_category = await client.post(
        "/api/v1/finances/categories",
        headers=owner_headers,
        json={"name": "Income", "type": "income"},
    )
    assert income_category.status_code == 201
    for payload in (
        {
            "account_id": account_id,
            "category_id": income_category.json()["id"],
            "type": "income",
            "amount": 1000,
            "transaction_date": "2026-07-01",
        },
        {
            "account_id": account_id,
            "category_id": category_id,
            "type": "expense",
            "amount": 600,
            "transaction_date": "2026-07-31",
        },
    ):
        response = await client.post(
            "/api/v1/finances/transactions",
            headers=owner_headers,
            json=payload,
        )
        assert response.status_code == 201, response.text
    budget = await client.post(
        "/api/v1/finances/budgets",
        headers=owner_headers,
        json={"category_id": category_id, "month": "2026-07", "amount": 500},
    )
    assert budget.status_code == 201
    recurring = await client.post(
        "/api/v1/finances/recurring",
        headers=owner_headers,
        json={
            "account_id": account_id,
            "category_id": category_id,
            "type": "expense",
            "amount": 40,
            "description": "Projection only until registered",
            "frequency": "daily",
            "start_date": "2026-07-15",
        },
    )
    assert recurring.status_code == 201
    registration = await client.post(
        f"/api/v1/finances/recurring/{recurring.json()['id']}/registrations",
        headers=owner_headers,
        json={"transaction_date": "2026-07-15"},
    )
    assert registration.status_code == 201

    other_account_id, other_category_id = await _create_account_and_category(client, other)
    other_transaction = await client.post(
        "/api/v1/finances/transactions",
        headers=_headers(other),
        json={
            "account_id": other_account_id,
            "category_id": other_category_id,
            "type": "expense",
            "amount": 9999,
            "transaction_date": "2026-07-31",
        },
    )
    assert other_transaction.status_code == 201

    response = await client.get(
        "/api/v1/finances/reports/monthly.xlsx",
        headers=owner_headers,
        params={"month": "2026-07"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument")
    assert response.headers["cache-control"] == "private, no-store"
    assert response.headers["content-disposition"].endswith('habitflow-monthly-report-2026-07.xlsx"')
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Spending by Category", "Budgets", "Six-Month Trends"]
    assert "Insights" not in workbook.sheetnames

    summary = workbook["Summary"]
    assert [cell.value for cell in summary[1]] == [
        "month", "period_start", "period_end", "total_income", "total_expenses", "net",
        "transaction_count", "income_transaction_count", "expense_transaction_count", "previous_month",
        "previous_period_start", "previous_period_end", "previous_total_income", "previous_total_expenses",
        "previous_net", "previous_transaction_count", "income_absolute_change", "income_percentage_change",
        "expenses_absolute_change", "expenses_percentage_change", "net_absolute_change", "net_percentage_change",
    ]
    assert summary["A2"].value == "2026-07"
    assert summary["D2"].value == 1000
    assert summary["E2"].value == 640
    assert summary["R2"].value is None
    assert summary["R2"].number_format == "0.00"
    assert summary.freeze_panes == "A2"
    assert summary.auto_filter.ref == "A1:V2"
    assert summary["A1"].font.bold is True

    spending = workbook["Spending by Category"]
    assert spending["B2"].value == "'=Formula category"
    assert spending["C2"].value == 640
    assert spending["E2"].number_format == "0.00"
    assert "9999" not in response.content.decode("latin1")

    budgets = workbook["Budgets"]
    assert budgets["D2"].value == 640
    assert budgets["H2"].value == 128.0
    assert budgets["H2"].number_format == "0.00"

    trends = workbook["Six-Month Trends"]
    assert trends.max_row == 7
    assert trends["A2"].value == "2026-02"
    assert trends["A7"].value == "2026-07"
    assert trends["J2"].value is None
    assert trends["J7"].value == 36.0
    assert trends["J7"].number_format == "0.00"


async def test_monthly_report_xlsx_empty_data_uses_app_month(client: AsyncClient, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(finances_service, "current_app_date", lambda: date(2026, 1, 22))
    token = await _register_and_login(client, "empty-report-export@example.com")
    response = await client.get(
        "/api/v1/finances/reports/monthly.xlsx",
        headers=_headers(token),
    )
    assert response.status_code == 200
    assert response.headers["content-disposition"].endswith('habitflow-monthly-report-2026-01.xlsx"')
    workbook = load_workbook(BytesIO(response.content))
    assert workbook["Summary"].max_row == 2
    assert workbook["Spending by Category"].max_row == 1
    assert workbook["Budgets"].max_row == 1
    assert workbook["Six-Month Trends"].max_row == 7


@pytest.mark.parametrize("month", ["2026-7", "2026-13", "0000-01"])
async def test_monthly_report_xlsx_rejects_invalid_month(client: AsyncClient, month: str) -> None:
    token = await _register_and_login(client, f"invalid-report-export-{month}@example.com")
    response = await client.get(
        "/api/v1/finances/reports/monthly.xlsx",
        headers=_headers(token),
        params={"month": month},
    )
    assert response.status_code == 422


async def test_monthly_report_xlsx_allows_future_month(client: AsyncClient) -> None:
    token = await _register_and_login(client, "future-report-export@example.com")
    response = await client.get(
        "/api/v1/finances/reports/monthly.xlsx",
        headers=_headers(token),
        params={"month": "2030-01"},
    )
    assert response.status_code == 200
    assert response.headers["content-disposition"].endswith('habitflow-monthly-report-2030-01.xlsx"')
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Spending by Category", "Budgets", "Six-Month Trends"]
    assert "Insights" not in workbook.sheetnames
    assert workbook["Summary"].max_row == 2


async def test_transaction_exports_filter_scope_and_sanitize_cells(client: AsyncClient) -> None:
    owner = await _register_and_login(client, "export-owner@example.com")
    other = await _register_and_login(client, "export-other@example.com")
    account_id, category_id = await _create_account_and_category(client, owner)
    other_account_id, other_category_id = await _create_account_and_category(client, other)

    created = await client.post(
        "/api/v1/finances/transactions",
        headers=_headers(owner),
        json={
            "account_id": account_id,
            "category_id": category_id,
            "type": "expense",
            "amount": 1200,
            "description": "=SUM(A1:A2)",
            "transaction_date": "2026-07-01",
        },
    )
    assert created.status_code == 201
    await client.post(
        "/api/v1/finances/transactions",
        headers=_headers(other),
        json={
            "account_id": other_account_id,
            "category_id": other_category_id,
            "type": "expense",
            "amount": 9999,
            "description": "Other user",
            "transaction_date": "2026-07-01",
        },
    )

    response = await client.get(
        "/api/v1/finances/exports/transactions.csv",
        headers=_headers(owner),
        params={"from": "2026-07-01", "to": "2026-07-31", "type": "expense"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert response.headers["cache-control"] == "private, no-store"
    assert response.headers["content-disposition"].endswith('habitflow-transactions-2026-07-01-to-2026-07-31.csv"')
    rows = _csv_rows(response)
    assert rows[0] == [
        "transaction_date", "type", "amount", "account_name", "category_name", "description",
        "transaction_id", "account_id", "category_id", "created_at",
    ]
    assert len(rows) == 2
    assert rows[1][2] == "1200"
    assert rows[1][4] == "'=Formula category"
    assert rows[1][5] == "'=SUM(A1:A2)"
    assert "9999" not in response.text

    workbook_response = await client.get(
        "/api/v1/finances/exports/transactions.xlsx",
        headers=_headers(owner),
        params={"from": "2026-07-01", "to": "2026-07-31", "sort": "asc"},
    )
    assert workbook_response.status_code == 200
    assert workbook_response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument")
    workbook = load_workbook(BytesIO(workbook_response.content))
    worksheet = workbook.active
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:J2"
    assert worksheet["A1"].font.bold is True
    assert worksheet["E2"].value == "'=Formula category"
    assert worksheet["F2"].value == "'=SUM(A1:A2)"


@pytest.mark.parametrize(
    "params",
    [
        {"from": "2026-07-01"},
        {"from": "2026-07-02", "to": "2026-07-01"},
        {"from": "2025-01-01", "to": "2026-01-02"},
    ],
)
async def test_transaction_export_rejects_invalid_ranges(client: AsyncClient, params: dict[str, str]) -> None:
    token = await _register_and_login(client, f"invalid-export-{len(params)}-{params['from']}@example.com")
    response = await client.get(
        "/api/v1/finances/exports/transactions.csv",
        headers=_headers(token),
        params=params,
    )
    assert response.status_code == 422


async def test_budget_exports_reuse_budget_progress(client: AsyncClient) -> None:
    token = await _register_and_login(client, "budget-export@example.com")
    account_id, category_id = await _create_account_and_category(client, token)
    headers = _headers(token)
    await client.post(
        "/api/v1/finances/transactions",
        headers=headers,
        json={
            "account_id": account_id,
            "category_id": category_id,
            "type": "expense",
            "amount": 600,
            "description": "Budget spending",
            "transaction_date": "2026-07-31",
        },
    )
    budget = await client.post(
        "/api/v1/finances/budgets",
        headers=headers,
        json={"category_id": category_id, "month": "2026-07", "amount": 500},
    )
    assert budget.status_code == 201

    csv_response = await client.get(
        "/api/v1/finances/exports/monthly-budgets.csv",
        headers=headers,
        params={"month": "2026-07"},
    )
    rows = _csv_rows(csv_response)
    assert rows[1][0:8] == ["2026-07", "'=Formula category", "500", "600", "0", "100", "1", "120.0"]
    assert rows[1][8] == "true"

    xlsx_response = await client.get(
        "/api/v1/finances/exports/monthly-budgets.xlsx",
        headers=headers,
        params={"month": "2026-07"},
    )
    workbook = load_workbook(BytesIO(xlsx_response.content))
    assert workbook.active["H2"].value == 120.0
    assert workbook.active["H2"].number_format == "0.00"


async def test_savings_exports_include_aggregated_contributions(client: AsyncClient) -> None:
    owner = await _register_and_login(client, "savings-export-owner@example.com")
    other = await _register_and_login(client, "savings-export-other@example.com")
    created = await client.post(
        "/api/v1/savings/goals",
        headers=_headers(owner),
        json={"name": "=Trip", "description": "=Unsafe", "target_amount": 1000},
    )
    goal_id = created.json()["id"]
    await client.post(
        f"/api/v1/savings/goals/{goal_id}/contributions",
        headers=_headers(owner),
        json={"amount": 400, "contribution_date": "2026-07-10"},
    )
    await client.post(
        "/api/v1/savings/goals",
        headers=_headers(other),
        json={"name": "Other goal", "target_amount": 10},
    )

    csv_response = await client.get("/api/v1/savings/exports/goals.csv", headers=_headers(owner))
    rows = _csv_rows(csv_response)
    assert len(rows) == 2
    assert rows[1][0] == "'=Trip"
    assert rows[1][1] == "'=Unsafe"
    assert rows[1][2:7] == ["1000", "400", "600", "40", "active"]

    xlsx_response = await client.get("/api/v1/savings/exports/goals.xlsx", headers=_headers(owner))
    workbook = load_workbook(BytesIO(xlsx_response.content))
    assert workbook.active["A2"].value == "'=Trip"
    assert workbook.active["D2"].value == 400


async def test_goal_contribution_exports_are_scoped_and_sanitize_notes(
    client: AsyncClient,
) -> None:
    owner = await _register_and_login(client, "contribution-export-owner@example.com")
    other = await _register_and_login(client, "contribution-export-other@example.com")
    created = await client.post(
        "/api/v1/savings/goals",
        headers=_headers(owner),
        json={"name": "Trip", "target_amount": 1000},
    )
    goal_id = created.json()["id"]
    await client.post(
        f"/api/v1/savings/goals/{goal_id}/contributions",
        headers=_headers(owner),
        json={"amount": 400, "note": "=Unsafe", "contribution_date": "2026-07-10"},
    )

    unauthenticated = await client.get(
        f"/api/v1/savings/exports/goals/{goal_id}/contributions.csv"
    )
    assert unauthenticated.status_code == 401

    other_user = await client.get(
        f"/api/v1/savings/exports/goals/{goal_id}/contributions.csv",
        headers=_headers(other),
    )
    assert other_user.status_code == 404

    csv_response = await client.get(
        f"/api/v1/savings/exports/goals/{goal_id}/contributions.csv",
        headers=_headers(owner),
    )
    assert csv_response.status_code == 200
    assert csv_response.headers["cache-control"] == "private, no-store"
    assert csv_response.headers["content-disposition"].endswith(
        f'habitflow-savings-contributions-{goal_id}-{current_app_date().isoformat()}.csv"'
    )
    rows = _csv_rows(csv_response)
    assert rows[0] == [
        "goal_name", "contribution_date", "amount", "note", "contribution_id",
        "goal_id", "created_at", "updated_at",
    ]
    assert rows[1][0:4] == ["Trip", "2026-07-10", "400", "'=Unsafe"]

    xlsx_response = await client.get(
        f"/api/v1/savings/exports/goals/{goal_id}/contributions.xlsx",
        headers=_headers(owner),
    )
    workbook = load_workbook(BytesIO(xlsx_response.content))
    assert workbook.active["D2"].value == "'=Unsafe"


async def test_empty_goal_contribution_exports_include_headers(client: AsyncClient) -> None:
    token = await _register_and_login(client, "empty-contribution-export@example.com")
    created = await client.post(
        "/api/v1/savings/goals",
        headers=_headers(token),
        json={"name": "Empty", "target_amount": 1000},
    )
    goal_id = created.json()["id"]

    csv_response = await client.get(
        f"/api/v1/savings/exports/goals/{goal_id}/contributions.csv",
        headers=_headers(token),
    )
    assert _csv_rows(csv_response) == [[
        "goal_name", "contribution_date", "amount", "note", "contribution_id",
        "goal_id", "created_at", "updated_at",
    ]]

    xlsx_response = await client.get(
        f"/api/v1/savings/exports/goals/{goal_id}/contributions.xlsx",
        headers=_headers(token),
    )
    workbook = load_workbook(BytesIO(xlsx_response.content))
    assert workbook.active.max_row == 1
